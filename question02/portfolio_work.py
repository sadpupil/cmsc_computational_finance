import numpy as np;
import math;
from openpyxl import load_workbook;
from openpyxl import Workbook;
from datetime import datetime;

class RawData:
    def __init__(self, date, price_spy, price_govt):
        self.date = date
        self.price_spy = price_spy
        self.price_govt = price_govt

class PortfolioValue:
    def __init__(self, portfolio_value):
        self.portofolio_value = portfolio_value

ratio = [0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1]

def read_data_from_xl(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    data_list = []
    
    for row in sheet.iter_rows(min_row = 3, values_only = True):
        date, price_spy, price_govt = row[1], row[2], row[3]
        date_formatted = date.strftime("%Y-%m-%d")
        data = RawData(date_formatted, price_spy, price_govt)  
        data_list.append(data) 
        # print(row[1], row[2], row[3], sep='')
    
    # for data in data_list:
    #     print(data.date, data.price_spy, data.price_govt)
    
    return data_list

def get_row_data(data, index, shares_spy, shares_govt):
    p_val = shares_spy * data.price_spy + shares_govt * data.price_govt
    if index == 0:
        return [data.date, round(data.price_spy, 2), round(data.price_govt, 2), round(p_val, 2), round(shares_spy, 2), round(shares_govt, 2)]
    else:
        return [data.date, round(data.price_spy, 2), round(data.price_govt, 2), round(p_val, 2), '', '']
        
    
def construct_portfolio(init_capital):
    data_list = read_data_from_xl('data\CMSC 5718 assignment 3 data.xlsx')
    
    init_price_spy = data_list[0].price_spy
    init_price_govt = data_list[0].price_govt
    
    end_date_2021 = '2021-12-31'
    end_date_2022 = '2022-12-30'
    end_date_2023 = '2023-12-29'
    
    # split data_list by year
    list_1, list_2, list_3 = [], [], []
    for data in data_list:
        if datetime.strptime(data.date, "%Y-%m-%d").date() <= datetime.strptime(end_date_2021, "%Y-%m-%d").date():
            list_1.append(data)
        elif datetime.strptime(data.date, "%Y-%m-%d").date() <= datetime.strptime(end_date_2022, "%Y-%m-%d").date():
            list_2.append(data)
        else:
            list_3.append(data)
    
    output_workbook = Workbook()
    # remove the default sheet
    output_workbook.remove(output_workbook.active)
    
    # with different equity/bond mix
    # store the portfolio value on 2020/12/31 and the value on 2023/12/29
    # for return and risk calculation
    portfolio_value_pairs = []
    
    # from 0%-100% to 100%-0%
    # store the no. of shares of spy and govt (2021, 2022, 2023)
    number_of_shares = []
    
    # ratio = [0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1]
    for r in ratio:
        sheet = output_workbook.create_sheet(str(r * 100) + '%' + '-' + str(100 - r * 100) + '%')
        sheet['A1'] = 'Date'
        sheet['B1'] = 'SPY Price'
        sheet['C1'] = 'GOVT Price'
        sheet['D1'] = 'Daily Portfolio Value'
        sheet['E1'] = 'Number of Shares (SPY)'
        sheet['F1'] = 'Number of Shares (GOVT)'
        
        weight_spy = r
        weight_govt = 1 - r
    
        shares_spy = init_capital * weight_spy / init_price_spy
        shares_govt = init_capital * weight_govt / init_price_govt
        shares_list = []
        shares_list.extend([round(shares_spy, 2), round(shares_govt, 2)])
        
        value_last = 0.0
    
        for index, data in enumerate(list_1):
            p_val = shares_spy * data.price_spy + shares_govt * data.price_govt
            sheet.append(get_row_data(data, index, shares_spy, shares_govt))
            value_last = p_val
    
        # re-balancing
        shares_spy = value_last * weight_spy / list_1[-1].price_spy
        shares_govt = value_last * weight_govt / list_1[-1].price_govt
        shares_list.extend([round(shares_spy, 2), round(shares_govt, 2)])
        for index, data in enumerate(list_2):
            p_val = shares_spy * data.price_spy + shares_govt * data.price_govt
            sheet.append(get_row_data(data, index, shares_spy, shares_govt))
            value_last = p_val
        
        # re-balancing
        shares_spy = value_last * weight_spy / list_2[-1].price_spy
        shares_govt = value_last * weight_govt / list_2[-1].price_govt
        shares_list.extend([round(shares_spy, 2), round(shares_govt, 2)])
        for index, data in enumerate(list_3):
            p_val = shares_spy * data.price_spy + shares_govt * data.price_govt
            sheet.append(get_row_data(data, index, shares_spy, shares_govt))
            
        value_start = sheet.cell(row=2, column=4).value
        value_end = sheet.cell(row=sheet.max_row, column=4).value
        # print('v1 = ', value_start, 'v2 = ', value_end)
        portfolio_value_pairs.append([value_start, value_end])
        
        number_of_shares.append(shares_list)

    save_filepath = 'data\daily portfolio value.xlsx'
    output_workbook.save(save_filepath)
    
    return number_of_shares, portfolio_value_pairs, data_list

def calculate_return(portfolio_value_pairs):
    # the range of year is 3
    n = 3
    
    portfolio_returns = []
    
    for values in portfolio_value_pairs:
        # print(values)
        start_value = values[0]
        end_value = values[1]
        rtn_value = (end_value - start_value) / start_value
        annualized_rtn = ((1 + rtn_value) ** (1 / n) - 1) * 100
        annualized_rtn = round(annualized_rtn, 2)
        portfolio_returns.append(annualized_rtn)
        # print('return: ', annualized_rtn, '%', sep='')
    
    return portfolio_returns
        
def calculate_risk_and_contribution(data_list):
    
    portfolio_risk = []
    risk_contribution_ratio = []
    
    for r in ratio:
        t = 252
        share_spy = r
        share_govt = 1 - r
        rtn_arr_spy = []
        rtn_arr_govt = []
        pre_price_spy = data_list[0].price_spy
        pre_price_govt = data_list[0].price_govt
        for index, data in enumerate(data_list):
            cur_price_spy = data.price_spy
            cur_price_govt = data.price_govt
            if index > 0:
                rtn_spy = (cur_price_spy / pre_price_spy) - 1
                rtn_govt = (cur_price_govt / pre_price_govt) - 1
                rtn_arr_spy.append(rtn_spy)
                rtn_arr_govt.append(rtn_govt)
            pre_price_spy = cur_price_spy
            pre_price_govt = cur_price_govt
            
        rtn_avg_spy = sum(rtn_arr_spy) / len(rtn_arr_spy)
        rtn_avg_govt = sum(rtn_arr_govt) / len(rtn_arr_govt)
    
        corr_spy_spy = 0.0
        corr_govt_govt = 0.0
        corr_spy_govt = 0.0
    
        # calculate the correlation of spy
        for data in rtn_arr_spy:
            corr_spy_spy += (data - rtn_avg_spy) ** 2
    
        corr_spy_spy = math.sqrt(corr_spy_spy / (len(rtn_arr_spy) - 1))
        # adjust with annualized factor
        corr_spy_spy *= math.sqrt(t)
    
        # calculate the correlation of govt
        for data in rtn_arr_govt:
            corr_govt_govt += (data - rtn_avg_govt) ** 2
    
        corr_govt_govt = math.sqrt(corr_govt_govt / (len(rtn_arr_govt) - 1))
        # adjust with annualized factor
        corr_govt_govt *= math.sqrt(t)
    
        # calculate the correlation of spy and govt
        for i in range(len(rtn_arr_spy)):
            corr_spy_govt += (rtn_arr_spy[i] - rtn_avg_spy) * (rtn_arr_govt[i] - rtn_avg_govt)
        corr_spy_govt *= (t / (len(rtn_arr_spy) - 1))
    
        corr_coeff = corr_spy_govt / (corr_spy_spy * corr_govt_govt)
    
        protfolio_variance = share_spy ** 2 * corr_spy_spy ** 2 + share_govt ** 2 * corr_govt_govt ** 2 + 2 * corr_coeff * share_spy * share_govt * corr_spy_spy * corr_govt_govt
        risk = math.sqrt(protfolio_variance)
        # print(portfolio_risk)
        portfolio_risk.append(round(risk  * 100, 2))
        
        # calculate risk contribution ratio
        contribution_spy = share_spy ** 2 * corr_spy_spy ** 2 + share_spy * share_govt * corr_coeff * corr_spy_spy * corr_govt_govt
        contribution_govt = share_govt ** 2 * corr_govt_govt ** 2 + share_spy * share_govt * corr_coeff * corr_spy_spy * corr_govt_govt
        
        contribution_spy /= risk
        contribution_govt /= risk
                
        contribution_ratio_spy = contribution_spy / risk
        contribution_ratio_govt = contribution_govt / risk
        
        risk_contribution_ratio.append([round(contribution_ratio_spy * 100, 2), round(contribution_ratio_govt * 100, 2)])
        
    return portfolio_risk, risk_contribution_ratio
        