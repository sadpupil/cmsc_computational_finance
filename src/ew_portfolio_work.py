import math;
import numpy as np;
from openpyxl import load_workbook;
from openpyxl import Workbook;
from datetime import datetime;
from portfolio_work import read_data_from_xl;

weight_spy = 0.3333
weight_govt = 0.3333
weight_gsg = 0.3333

def get_row_data(data, index, shares_spy, shares_govt, shares_gsg):
    p_val = shares_spy * data.price_spy + shares_govt * data.price_govt + shares_gsg * data.price_gsg
    if index == 0:
        return [data.date, round(data.price_spy, 2), round(data.price_govt, 2), round(data.price_gsg, 2), round(p_val, 2), round(shares_spy, 2), round(shares_govt, 2), round(shares_gsg, 2)]
    else:
        return [data.date, round(data.price_spy, 2), round(data.price_govt, 2), round(data.price_gsg, 2), round(p_val, 2), '', '']

def construct_ew_portfolio(initial_capital):
    data_list = read_data_from_xl('src_data\CMSC 5718 assignment 3 data.xlsx')
    
    end_date_2021 = '2021-12-31'
    end_date_2022 = '2022-12-30'
    year_list = ['2021', '2022', '2023']
    
    # split data_list by year
    list_1, list_2, list_3 = [], [], []
    for data in data_list:
        if datetime.strptime(data.date, "%Y-%m-%d").date() <= datetime.strptime(end_date_2021, "%Y-%m-%d").date():
            list_1.append(data)
        elif datetime.strptime(data.date, "%Y-%m-%d").date() <= datetime.strptime(end_date_2022, "%Y-%m-%d").date():
            list_2.append(data)
        else:
            list_3.append(data)
    
    total_list = []        
    total_list.extend([list_1, list_2, list_3])
    
    daily_value = [[], [], []]
    number_of_shares = []
    
    output_workbook = Workbook()
    # remove the default sheet
    output_workbook.remove(output_workbook.active)
    
    # construc the portfolio of each year separatly
    for index, list in enumerate(total_list):
        sheet = output_workbook.create_sheet('year ' + year_list[index])
        sheet['A1'] = 'Date'
        sheet['B1'] = 'SPY Price'
        sheet['C1'] = 'GOVT Price'
        sheet['D1'] = 'GSG Price'
        sheet['E1'] = 'Daily Portfolio Value'
        sheet['F1'] = 'Number of Shares (SPY)'
        sheet['G1'] = 'Number of Shares (GOVT)'
        sheet['H1'] = 'Number of Shares (GSG)'
        
        init_price_spy = list[0].price_spy
        init_price_govt = list[0].price_govt
        init_price_gsg = list[0].price_gsg
        
        shares_spy = initial_capital * weight_spy / init_price_spy
        shares_govt = initial_capital * weight_govt / init_price_govt
        shares_gsg = initial_capital * weight_gsg / init_price_gsg
        
        number_of_shares.append([round(shares_spy, 2), round(shares_govt, 2), round(shares_gsg, 2)])
        
        for j, data in enumerate(list):
            # daily value of the portfolio
            val = shares_spy * data.price_spy + shares_govt * data.price_govt + shares_gsg * data.price_gsg
            daily_value[index].append(val)
            sheet.append(get_row_data(data, j, shares_spy, shares_govt, shares_gsg))
            
    save_filepath = 'result_data\equal weight portfolio value.xlsx'
    output_workbook.save(save_filepath)
    
    # return list_2 to be question 3-b's input    
    return total_list, daily_value, number_of_shares, list_2, list_3

def calculate_ew_return_and_risk(daily_value):
    
    result_list = [[], [], []]
    
    for index, elem in enumerate(daily_value):
        start_price = elem[0]
        end_price = elem[-1]
        
        # !!
        portfolio_return = (end_price - start_price) / start_price
        
        # calculate the standard deviation of return
        rtn_list = []
        pre_value = elem[0]
        for j, val in enumerate(elem):
            if j > 0:
                # Rki
                rtn_list.append((val / pre_value) - 1)  
            pre_value = val
                
        avg_return = sum(rtn_list) / len(rtn_list)
        
        # !!
        portfolio_stand_dev = 0
        for r in rtn_list:
            portfolio_stand_dev += (r - avg_return) ** 2
        
        portfolio_stand_dev /= (len(rtn_list) - 1)
        # portfolio_stand_dev = math.sqrt(portfolio_stand_dev)
        portfolio_stand_dev = math.sqrt(portfolio_stand_dev) * math.sqrt(252)
        
        ratio = portfolio_return / portfolio_stand_dev
        
        result_list[index].append(round(portfolio_return * 100, 2))
        result_list[index].append(round(portfolio_stand_dev * 100, 2))
        result_list[index].append(round(ratio, 2))
    
    return result_list

# calculate the covariance matrix's elements
def calculate_covariance_matrix_elem(rtn_list_i, rtn_list_j, rtn_avg_i, rtn_avg_j):
    ans = 0
    for k in range(len(rtn_list_i)):
        rtn_value_i = rtn_list_i[k]
        rtn_value_j = rtn_list_j[k]
        ans += ((rtn_value_i - rtn_avg_i) * (rtn_value_j - rtn_avg_j))
    
    ans *= (252 / (len(rtn_list_i) - 1))
    # return ans / len(rtn_list_i)
    return ans
    

def calculate_ew_risk_contribution_ratio(total_list):
    # calculate by year
    # following the sequence of 2021 -> 2022 -> 2023
    result_list = []
    for i, list in enumerate(total_list):
        rtn_spy = []
        rtn_govt = []
        rtn_gsg = []
        
        # calculate the everyday return
        # using the formed list to calculate the avg
        pre_value_spy = list[0].price_spy
        pre_value_govt = list[0].price_govt
        pre_value_gsg = list[0].price_gsg
        for j, data in enumerate(list):
            if j > 0:
                rtn_spy.append(data.price_spy / pre_value_spy - 1)
                rtn_govt.append(data.price_govt / pre_value_govt - 1)
                rtn_gsg.append(data.price_gsg / pre_value_gsg - 1)
            pre_value_spy = data.price_spy
            pre_value_govt = data.price_govt
            pre_value_gsg = data.price_gsg
        
        # calculate th avg of return for each asset
        avg_rtn_spy = sum(rtn_spy) / len(rtn_spy)
        avg_rtn_govt = sum(rtn_govt) / len(rtn_govt)
        avg_rtn_gsg = sum(rtn_gsg) / len(rtn_gsg)
        
        # build the covariance matrix
        # first calculate each sigma within the matrix
        sig_1_1 = calculate_covariance_matrix_elem(rtn_spy, rtn_spy, avg_rtn_spy, avg_rtn_spy)
        sig_1_2 = calculate_covariance_matrix_elem(rtn_spy, rtn_govt, avg_rtn_spy, avg_rtn_govt)
        sig_1_3 = calculate_covariance_matrix_elem(rtn_spy, rtn_gsg, avg_rtn_spy, avg_rtn_gsg)
        
        sig_2_1 = calculate_covariance_matrix_elem(rtn_govt, rtn_spy, avg_rtn_govt, avg_rtn_spy)
        sig_2_2 = calculate_covariance_matrix_elem(rtn_govt, rtn_govt, avg_rtn_govt, avg_rtn_govt)
        sig_2_3 = calculate_covariance_matrix_elem(rtn_govt, rtn_gsg, avg_rtn_govt, avg_rtn_gsg)
        
        sig_3_1 = calculate_covariance_matrix_elem(rtn_gsg, rtn_spy, avg_rtn_gsg, avg_rtn_spy)
        sig_3_2 = calculate_covariance_matrix_elem(rtn_gsg, rtn_govt, avg_rtn_gsg, avg_rtn_govt)
        sig_3_3 = calculate_covariance_matrix_elem(rtn_gsg, rtn_gsg, avg_rtn_gsg, avg_rtn_gsg)
        
        covar_matrix = np.array([
            [sig_1_1, sig_1_2, sig_1_3],
            [sig_2_1, sig_2_2, sig_2_3],
            [sig_3_1, sig_3_2, sig_3_3],
        ])
        
        # print(covar_matrix)
        
        w = np.array([
            [weight_spy], 
            [weight_govt], 
            [weight_gsg]
        ])
        
        MR = np.dot(covar_matrix, w)
        # print(MR, MR[0], MR[1], MR[2])
        sig_p = np.dot(np.dot(np.transpose(w), covar_matrix), w)
        sig_p = math.sqrt(sig_p)
        # print(sig_p)
        
        mr = [MR[0][0] / sig_p, MR[1][0] / sig_p, MR[2][0] / sig_p]
        tr = [mr[0] * weight_spy / sig_p, mr[1] * weight_govt / sig_p, mr[2] * weight_gsg / sig_p]
        contribution_ratio = [round(tr[0] * 100, 2), round(tr[1] * 100, 2), round(tr[2] * 100, 2)]
        result_list.append(contribution_ratio) 
        
    return result_list
    
    