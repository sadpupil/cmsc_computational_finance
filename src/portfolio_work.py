import math;
import numpy as np;
from openpyxl import load_workbook;
from openpyxl import Workbook;
from datetime import datetime;

class RawData:
    def __init__(self, date, price_spy, price_govt, price_gsg):
        self.date = date
        self.price_spy = price_spy
        self.price_govt = price_govt
        self.price_gsg = price_gsg
        
class PortfolioData:
    def __init__(self, date, price_spy, price_govt, daily_value):
        self.date = date
        self.price_spy = price_spy
        self.price_govt = price_govt
        self.daily_value = daily_value

ratio = [0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1]

def read_data_from_xl(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    data_list = []
    
    for row in sheet.iter_rows(min_row = 3, values_only = True):
        date, price_spy, price_govt, price_gsg = row[1], row[2], row[3], row[4]
        date_formatted = date.strftime("%Y-%m-%d")
        data = RawData(date_formatted, price_spy, price_govt, price_gsg)  
        data_list.append(data) 
        # print(row[1], row[2], row[3], sep='')
    
    # for data in data_list:
    #     print(data.date, data.price_spy, data.price_govt)
    
    return data_list

def get_row_data(data, index, shares_spy, shares_govt):
    p_val = shares_spy * data.price_spy + shares_govt * data.price_govt
    if index == 0:
        return [data.date, round(data.price_spy, 2), round(data.price_govt, 2), round(p_val, 2), round(shares_spy, 2), round(shares_govt, 2)]
    else:
        return [data.date, round(data.price_spy, 2), round(data.price_govt, 2), round(p_val, 2), '', '']
        
    
def construct_portfolio(init_capital):
    data_list = read_data_from_xl('src_data\CMSC 5718 assignment 3 data.xlsx')
    
    init_price_spy = data_list[0].price_spy
    init_price_govt = data_list[0].price_govt
    
    end_date_2021 = '2021-12-31'
    end_date_2022 = '2022-12-30'
    end_date_2023 = '2023-12-29'
    
    # split data_list by year
    list_1, list_2, list_3 = [], [], []
    for data in data_list:
        if datetime.strptime(data.date, "%Y-%m-%d").date() <= datetime.strptime(end_date_2021, "%Y-%m-%d").date():
            list_1.append(data)
        elif datetime.strptime(data.date, "%Y-%m-%d").date() <= datetime.strptime(end_date_2022, "%Y-%m-%d").date():
            list_2.append(data)
        else:
            list_3.append(data)
    
    output_workbook = Workbook()
    # remove the default sheet
    output_workbook.remove(output_workbook.active)
    
    # from 0%-100% to 100%-0%
    # store the no. of shares of spy and govt (2021, 2022, 2023)
    number_of_shares = []
    
    price_value_list = []
    
    for r in ratio:
        sheet = output_workbook.create_sheet(str(r * 100) + '%' + '-' + str(100 - r * 100) + '%')
        sheet['A1'] = 'Date'
        sheet['B1'] = 'SPY Price'
        sheet['C1'] = 'GOVT Price'
        sheet['D1'] = 'Daily Portfolio Value'
        sheet['E1'] = 'Number of Shares (SPY)'
        sheet['F1'] = 'Number of Shares (GOVT)'
        
        weight_spy = r
        weight_govt = 1 - r
    
        shares_spy = init_capital * weight_spy / init_price_spy
        shares_govt = init_capital * weight_govt / init_price_govt
        shares_list = []
        shares_list.extend([round(shares_spy, 2), round(shares_govt, 2)])
        
        price_value_container = []
        
        value_last = 0.0
    
        for index, data in enumerate(list_1):
            p_val = shares_spy * data.price_spy + shares_govt * data.price_govt
            price_value_container.append(PortfolioData(data.date, data.price_spy, data.price_govt, p_val))
            sheet.append(get_row_data(data, index, shares_spy, shares_govt))
            value_last = p_val
    
        # re-balancing
        shares_spy = value_last * weight_spy / list_1[-1].price_spy
        shares_govt = value_last * weight_govt / list_1[-1].price_govt
        shares_list.extend([round(shares_spy, 2), round(shares_govt, 2)])
        for index, data in enumerate(list_2):
            p_val = shares_spy * data.price_spy + shares_govt * data.price_govt
            price_value_container.append(PortfolioData(data.date, data.price_spy, data.price_govt, p_val))
            sheet.append(get_row_data(data, index, shares_spy, shares_govt))
            value_last = p_val
        
        # re-balancing
        shares_spy = value_last * weight_spy / list_2[-1].price_spy
        shares_govt = value_last * weight_govt / list_2[-1].price_govt
        shares_list.extend([round(shares_spy, 2), round(shares_govt, 2)])
        for index, data in enumerate(list_3):
            p_val = shares_spy * data.price_spy + shares_govt * data.price_govt
            price_value_container.append(PortfolioData(data.date, data.price_spy, data.price_govt, p_val))
            sheet.append(get_row_data(data, index, shares_spy, shares_govt))
        
        number_of_shares.append(shares_list)
        price_value_list.append(price_value_container)

    save_filepath = 'result_data\daily portfolio value.xlsx'
    output_workbook.save(save_filepath)
    
    return price_value_list, number_of_shares
    
    portfolio_risk = calculate_risk_dailyValue(daily_value_set)
    risk_contribution_ratio = []
    
    for j, r in enumerate(ratio):
        t = 252
        share_spy = r
        share_govt = 1 - r
        rtn_arr_spy = []
        rtn_arr_govt = []
        pre_price_spy = data_list[0].price_spy
        pre_price_govt = data_list[0].price_govt
        for index, data in enumerate(data_list):
            cur_price_spy = data.price_spy
            cur_price_govt = data.price_govt
            if index > 0:
                rtn_spy = (cur_price_spy / pre_price_spy) - 1
                rtn_govt = (cur_price_govt / pre_price_govt) - 1
                rtn_arr_spy.append(rtn_spy)
                rtn_arr_govt.append(rtn_govt)
            pre_price_spy = cur_price_spy
            pre_price_govt = cur_price_govt
            
        rtn_avg_spy = sum(rtn_arr_spy) / len(rtn_arr_spy)
        rtn_avg_govt = sum(rtn_arr_govt) / len(rtn_arr_govt)
    
        corr_spy_spy = 0.0
        corr_govt_govt = 0.0
        corr_spy_govt = 0.0
    
        # calculate the correlation of spy
        for data in rtn_arr_spy:
            corr_spy_spy += (data - rtn_avg_spy) ** 2
    
        corr_spy_spy = math.sqrt(corr_spy_spy / (len(rtn_arr_spy) - 1))
        # adjust with annualized factor
        corr_spy_spy *= math.sqrt(t)
    
        # calculate the correlation of govt
        for data in rtn_arr_govt:
            corr_govt_govt += (data - rtn_avg_govt) ** 2
    
        corr_govt_govt = math.sqrt(corr_govt_govt / (len(rtn_arr_govt) - 1))
        # adjust with annualized factor
        corr_govt_govt *= math.sqrt(t)
    
        # calculate the correlation of spy and govt
        for i in range(len(rtn_arr_spy)):
            corr_spy_govt += (rtn_arr_spy[i] - rtn_avg_spy) * (rtn_arr_govt[i] - rtn_avg_govt)
        corr_spy_govt *= (t / (len(rtn_arr_spy) - 1))
    
        corr_coeff = corr_spy_govt / (corr_spy_spy * corr_govt_govt)
    
        # protfolio_variance = share_spy ** 2 * corr_spy_spy ** 2 + share_govt ** 2 * corr_govt_govt ** 2 + 2 * corr_coeff * share_spy * share_govt * corr_spy_spy * corr_govt_govt
        # risk = math.sqrt(protfolio_variance)
        risk = portfolio_risk[j]
        # portfolio_risk.append(round(risk * 100, 2))
        portfolio_risk[j] = round(risk * 100, 2)
        
        # calculate risk contribution ratio
        contribution_spy = share_spy ** 2 * corr_spy_spy ** 2 + share_spy * share_govt * corr_coeff * corr_spy_spy * corr_govt_govt
        contribution_govt = share_govt ** 2 * corr_govt_govt ** 2 + share_spy * share_govt * corr_coeff * corr_spy_spy * corr_govt_govt
        
        contribution_spy /= risk
        contribution_govt /= risk
                
        contribution_ratio_spy = contribution_spy / risk
        contribution_ratio_govt = contribution_govt / risk
        
        risk_contribution_ratio.append([round(contribution_ratio_spy * 100, 2), round(contribution_ratio_govt * 100, 2)])
        
    return portfolio_risk, risk_contribution_ratio

def calculate_return_and_risk(price_value_list):
    # calculate return
    rtn_list = []
    for price_value in price_value_list:
        start_price = price_value[0].daily_value
        end_price = price_value[-1].daily_value
        tmp = (end_price - start_price) / start_price
        rtn_val = (1 + tmp) ** (1 / 3) - 1
        rtn_list.append(rtn_val) 
        # print(price_value[0], price_value[-1], sep='')
    
    # calculate annualized standard deviation
    stan_dev_list = []
    for price_value in price_value_list:
        r = []
        pre_value = price_value[0].daily_value
        
        for i, elem in enumerate(price_value):
            if i > 0:
                r.append(elem.daily_value / pre_value - 1)
            pre_value = elem.daily_value
        
        avg_r = sum(r) / len(r)
        # the standard deviation without annualization factor
        sig = 0
        for elem in r:
            sig += (elem - avg_r) ** 2
        sig /= (len(r) - 1)
        sig = math.sqrt(sig)
        
        # multiply with the square root of 252
        sig *= math.sqrt(252)
        stan_dev_list.append(sig)
        
    return rtn_list, stan_dev_list

# each element in the data list is a PortfolioData object
def calculate_covariance_matric_elem(mark_1, mark_2, data_list):
    rtn_list_1 = []
    rtn_list_2 = []
    
    if mark_1 == 'spy':
        pre = data_list[0].price_spy
        for i, elem in enumerate(data_list):
            if i > 0:
                tmp = (elem.price_spy / pre) - 1
                rtn_list_1.append(tmp)
            pre = elem.price_spy
    elif mark_1 == 'govt':
        pre = data_list[0].price_govt
        for i, elem in enumerate(data_list):
            if i > 0:
                tmp = (elem.price_govt / pre) - 1
                rtn_list_1.append(tmp)
            pre = elem.price_govt
    
    if mark_2 == 'spy':
        pre = data_list[0].price_spy
        for i, elem in enumerate(data_list):
            if i > 0:
                tmp = (elem.price_spy / pre) - 1
                rtn_list_2.append(tmp)
            pre = elem.price_spy
    elif mark_2 == 'govt':
        pre = data_list[0].price_govt
        for i, elem in enumerate(data_list):
            if i > 0:
                tmp = (elem.price_govt / pre) - 1
                rtn_list_2.append(tmp)
            pre = elem.price_govt 
            
    avg_rtn_1 = sum(rtn_list_1) / len(rtn_list_1)
    avg_rtn_2 = sum(rtn_list_2) / len(rtn_list_2)
    
    sig_1_2 = 0
    for j, r in enumerate(rtn_list_1):
        sig_1_2 += (rtn_list_1[j] - avg_rtn_1) * (rtn_list_2[j] - avg_rtn_2)
    
    # multiply annualized factor
    sig_1_2 *= (252 / (len(rtn_list_1) - 1))
    return sig_1_2

def contribution_ratio(weight_spy, weight_govt, price_value):
    w = np.array([
        [weight_spy],
        [weight_govt]
    ])
    
    sig_1_1 = calculate_covariance_matric_elem('spy', 'spy', price_value)
    sig_1_2 = calculate_covariance_matric_elem('spy', 'govt', price_value)
    sig_2_1 = sig_1_2
    sig_2_2 = calculate_covariance_matric_elem('govt', 'govt', price_value)
    
    covar_matrix = np.array([
        [sig_1_1, sig_1_2],
        [sig_2_1, sig_2_2]
    ])
    
    MR = np.dot(covar_matrix, w)
    sig_p_square = np.dot(np.dot(np.transpose(w), covar_matrix), w)
    sig_p = math.sqrt(sig_p_square)
    mr = [MR[0][0] / sig_p, MR[1][0] / sig_p]
    tr = [mr[0] * weight_spy / sig_p, mr[1] * weight_govt / sig_p]
    contribution_ratio = [round(tr[0] * 100, 2), round(tr[1] * 100, 2)]
    return contribution_ratio

def calculate_risk_contribution_ratio(price_value_list):
    risk_contribution_ratio = []
    for i, price_value in enumerate(price_value_list):
        risk_contribution_ratio.append(contribution_ratio(ratio[i], 1 - ratio[i], price_value))
    return risk_contribution_ratio